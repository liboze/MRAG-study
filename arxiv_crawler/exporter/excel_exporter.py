"""
Excel 导出模块
将论文元数据导出为英文版和中文翻译版两个 Excel 文件。
包含自动列宽调整、标题行加粗等格式化功能。
"""

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from tqdm import tqdm  # type: ignore

try:
    from ..translator.base import BaseTranslator  # 作为包导入时使用相对导入
except ImportError:
    from translator.base import BaseTranslator  # type: ignore  # 直接运行时使用绝对导入

logger = logging.getLogger(__name__)

# 英文列标题
EN_COLUMNS = [
    "Crawl Time",
    "Search Keywords",
    "Published Date",
    "Title",
    "Authors",
    "Affiliations",
    "Abstract",
    "arXiv ID",
    "PDF Link",
]

# 中文列标题
ZH_COLUMNS = [
    "爬取时间",
    "搜索关键词",
    "论文发表时间",
    "标题",
    "作者",
    "作者机构",
    "摘要",
    "arXiv ID",
    "PDF链接",
]

# 数据字段名（与列标题对应）
FIELD_KEYS = [
    "crawl_time",
    "keywords",
    "published_date",
    "title",
    "authors",
    "affiliations",
    "abstract",
    "arxiv_id",
    "pdf_link",
]

# 需要翻译的字段（其余字段保持原文）
TRANSLATABLE_FIELDS = {"title", "affiliations", "abstract"}

# 标题行背景颜色（深蓝色）
HEADER_FILL_COLOR = "1F4E79"
# 标题行字体颜色（白色）
HEADER_FONT_COLOR = "FFFFFF"


def _auto_adjust_column_widths(ws, max_width: int = 80) -> None:
    """
    自动调整工作表所有列的宽度，根据内容长度计算最优宽度。

    :param ws: openpyxl Worksheet 对象
    :param max_width: 最大列宽（字符数），默认 80
    """
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            if cell.value:
                # 中文字符约占 2 个字符宽度
                cell_str = str(cell.value)
                char_width = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                max_length = max(max_length, char_width)
        # 设置列宽：内容宽度 + 2 个字符边距，不超过 max_width
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 10)


def _format_header_row(ws, columns: List[str]) -> None:
    """
    格式化标题行：写入列名并应用加粗、背景色、白色字体、居中对齐。

    :param ws: openpyxl Worksheet 对象
    :param columns: 列标题列表
    """
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill(
        start_color=HEADER_FILL_COLOR,
        end_color=HEADER_FILL_COLOR,
        fill_type="solid",
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 固定标题行高度
    ws.row_dimensions[1].height = 20


def _write_data_rows(ws, papers: List[Dict[str, Any]], field_keys: List[str]) -> None:
    """
    将论文数据写入工作表，从第 2 行开始。

    :param ws: openpyxl Worksheet 对象
    :param papers: 论文数据列表
    :param field_keys: 字段名列表（与列顺序对应）
    """
    data_font = Font(size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    no_wrap_align = Alignment(vertical="top", wrap_text=False)

    # 需要自动换行的字段（长文本）
    wrap_fields = {"title", "abstract", "affiliations", "authors"}

    for row_idx, paper in enumerate(papers, start=2):
        for col_idx, field in enumerate(field_keys, start=1):
            value = paper.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = wrap_align if field in wrap_fields else no_wrap_align

        # 摘要行设置较高行高
        ws.row_dimensions[row_idx].height = 60


def _freeze_header_row(ws) -> None:
    """冻结第一行（标题行），方便滚动时始终可见。"""
    ws.freeze_panes = "A2"


class ExcelExporter:
    """
    将论文数据导出为格式化的 Excel 文件。
    生成英文原版和中文翻译版两个文件。
    """

    def __init__(
        self,
        output_dir: str,
        translator: Optional[BaseTranslator] = None,
    ) -> None:
        """
        初始化导出器。

        :param output_dir: 输出目录（会自动创建子目录结构）
        :param translator: 翻译器实例（为 None 时跳过中文版生成）
        """
        self.output_dir = output_dir
        self.translator = translator

    def export(
        self,
        papers: List[Dict[str, Any]],
        timestamp: Optional[datetime] = None,
    ) -> Dict[str, str]:
        """
        将论文数据导出为英文和中文两个 Excel 文件，并保存原始 JSON。

        :param papers: 论文数据列表
        :param timestamp: 文件名中使用的时间戳（默认为当前时间）
        :return: {"en": 英文文件路径, "zh": 中文文件路径, "json": JSON文件路径}
        """
        if not papers:
            logger.warning("没有论文数据可导出")
            return {}

        ts = timestamp or datetime.now()
        ts_str = ts.strftime("%Y%m%d_%H%M%S")
        date_str = ts.strftime("%Y-%m-%d")

        # 创建输出目录结构
        date_dir = os.path.join(self.output_dir, date_str)
        raw_dir = os.path.join(date_dir, "raw")
        os.makedirs(raw_dir, exist_ok=True)

        result_paths: Dict[str, str] = {}

        # 保存原始 JSON 数据
        json_path = os.path.join(raw_dir, f"arxiv_papers_raw_{ts_str}.json")
        self._save_json(papers, json_path)
        result_paths["json"] = json_path

        # 导出英文版 Excel
        en_path = os.path.join(date_dir, f"arxiv_papers_en_{ts_str}.xlsx")
        self._export_en(papers, en_path)
        result_paths["en"] = en_path
        logger.info("英文版 Excel 已保存：%s", en_path)

        # 导出中文翻译版 Excel
        if self.translator is not None:
            zh_path = os.path.join(date_dir, f"arxiv_papers_zh_{ts_str}.xlsx")
            self._export_zh(papers, zh_path)
            result_paths["zh"] = zh_path
            logger.info("中文版 Excel 已保存：%s", zh_path)
        else:
            logger.info("未配置翻译器，跳过中文版导出")

        return result_paths

    def _save_json(self, papers: List[Dict[str, Any]], path: str) -> None:
        """保存原始论文数据为 JSON 文件。"""
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(papers, f, ensure_ascii=False, indent=2)
            logger.info("原始 JSON 已保存：%s（共 %d 条）", path, len(papers))
        except OSError as exc:
            logger.error("保存 JSON 文件失败：%s", exc)

    def _export_en(self, papers: List[Dict[str, Any]], path: str) -> None:
        """
        导出英文原版 Excel 文件。

        :param papers: 论文数据列表
        :param path: 输出文件路径
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Papers (EN)"

        _format_header_row(ws, EN_COLUMNS)
        _write_data_rows(ws, papers, FIELD_KEYS)
        _freeze_header_row(ws)
        _auto_adjust_column_widths(ws)

        try:
            wb.save(path)
        except OSError as exc:
            logger.error("保存英文 Excel 失败：%s", exc)
            raise

    def _export_zh(self, papers: List[Dict[str, Any]], path: str) -> None:
        """
        翻译论文数据并导出中文版 Excel 文件。

        :param papers: 论文数据列表（英文原版）
        :param path: 输出文件路径
        """
        logger.info("开始翻译 %d 篇论文…", len(papers))

        # 批量翻译各字段，逐字段处理以便显示进度
        translated_papers: List[Dict[str, Any]] = [dict(p) for p in papers]

        for field in TRANSLATABLE_FIELDS:
            texts = [p.get(field, "") for p in papers]
            logger.info("翻译字段 [%s]（共 %d 条）…", field, len(texts))

            # 区分需要翻译的文本和占位符（"未提供"），只翻译有实际内容的条目
            indices_to_translate = [
                i for i, t in enumerate(texts) if t and t != "未提供"
            ]
            texts_to_translate = [texts[i] for i in indices_to_translate]

            translated_texts: List[str] = list(texts)  # 先复制原文

            if texts_to_translate:
                try:
                    # 使用批量翻译接口提升效率（OpenAI 会合并为单次 API 调用）
                    batch_results = self.translator.translate_batch(  # type: ignore
                        tqdm(texts_to_translate, desc=f"翻译 {field}", unit="条")
                        if len(texts_to_translate) > 1
                        else texts_to_translate
                    )
                    for idx, translated in zip(indices_to_translate, batch_results):
                        translated_texts[idx] = translated or texts[idx]
                except Exception as exc:  # pylint: disable=broad-except
                    logger.warning("批量翻译失败，字段=%s，原因=%s", field, exc)
                    # 回退到逐条翻译
                    for idx in tqdm(indices_to_translate, desc=f"翻译 {field}(重试)", unit="条"):
                        try:
                            result = self.translator.translate(texts[idx])  # type: ignore
                            translated_texts[idx] = result or texts[idx]
                        except Exception as inner_exc:  # pylint: disable=broad-except
                            logger.warning("单条翻译失败，保留原文。原因=%s", inner_exc)

            for i, trans_text in enumerate(translated_texts):
                translated_papers[i][field] = trans_text

        wb = Workbook()
        ws = wb.active
        ws.title = "论文列表（中文）"

        _format_header_row(ws, ZH_COLUMNS)
        _write_data_rows(ws, translated_papers, FIELD_KEYS)
        _freeze_header_row(ws)
        _auto_adjust_column_widths(ws)

        try:
            wb.save(path)
        except OSError as exc:
            logger.error("保存中文 Excel 失败：%s", exc)
            raise
